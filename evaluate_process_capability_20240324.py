#!/usr/bin/env python
# coding: utf-8

# In[2]:


import pandas as pd
import matplotlib.pyplot as plt
import japanize_matplotlib
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import tkinter as tk
from tkinter import filedialog
import sys


# In[3]:


# コピー元のパス・シート指定
# Tkインターフェイスを初期化（GUIウィンドウは表示されない）
root = tk.Tk()
root.withdraw()

# ファイル選択ダイアログを開く
file_path = filedialog.askopenfilename(
    title="ファイルを選択してください",
    filetypes=[("Excelファイル", "*.xlsx;*.xls")]
)

# Excelファイルから全てのシート名を取得
xls = pd.ExcelFile(file_path)
sheet_name_list = xls.sheet_names


# In[4]:


# xlsxファイルの読み込みとデータ整理
#元データ
original_df = pd.read_excel(file_path,sheet_name=sheet_name_list, skiprows=7, \
skipfooter=45, header=[0,1,2,3], index_col = 1)

#dictをリスト化&要らない列を削除
original_dfs = [original_df[sheet_name_list[sheet_number]].drop(('Unnamed: 0_level_0',\
'Unnamed: 0_level_1', 'Unnamed: 0_level_2', 'Unnamed: 0_level_3'), axis=1) \
for sheet_number in range(len(sheet_name_list))]

#df内の空白列を削除
for sheet_number in range(len(sheet_name_list)):
    original_dfs[sheet_number].dropna(how="all",axis=1, inplace=True)

#Multiple columnsをmargeしてoriginal_dfsを再定義
for sheet_number in range(len(sheet_name_list)):
    columns_integer = "Dim"
    columns_marge = [None] * len(original_dfs[sheet_number].columns)  # 47に相当するカラム数に合わせてリストを初期化

    for i in range(len(original_dfs[sheet_number].columns)):
        for j in range(4):
            if  "Unnamed" not in str(original_dfs[sheet_number].columns[i][j]) :
                columns_integer += "_" + str(original_dfs[sheet_number].columns[i][j])
        columns_marge[i] = columns_integer
        columns_integer = "Dim"

    for i in range(len(columns_marge)):
        columns_marge[i] = columns_marge[i].replace("/", "per")
        columns_marge[i] = columns_marge[i].replace(".", "")

    columns_marge[46]=columns_marge[46].replace(columns_marge[46],"Weight") #最後だけ手動で置き換え・・・


    original_dfs[sheet_number] = pd.DataFrame(original_dfs[sheet_number].values, \
    index = original_dfs[sheet_number].index, columns = columns_marge)


# In[5]:


#sampleデータのdf
sample_dfs = [original_dfs[sheet_number].drop(['Nominal', 'Up Tol.', 'Low Tol.']) for sheet_number in range(len(sheet_name_list))]

#寸法と公差のdf
dimension_dfs = [original_dfs[sheet_number].copy().head(3) for sheet_number in range(len(sheet_name_list))]


# In[6]:


# 統計量等のdf
statistics_dfs = []

for sdf, ddf in zip(sample_dfs, dimension_dfs):
    stats_df = pd.DataFrame()
    stats_df["Average"] = sdf.mean() #n30平均値
    stats_df["Max_sample"] = sdf.max() #n30最大値
    stats_df["Min_sample"] = sdf.min() #n30最小値
    stats_df["Sigma"] = sdf.std() #n30標準偏差
    stats_df["Upper_Limit"] = ddf.loc["Nominal"] + ddf.loc["Up Tol."] #最大寸法
    stats_df["Lower_Limit"] = ddf.loc["Nominal"] + ddf.loc["Low Tol."] #最小寸法

    stats_df["Center_Line"] = (stats_df["Upper_Limit"] + stats_df["Lower_Limit"])/2 #中央寸法
    stats_df["Process_Capability_Index"] = (stats_df["Upper_Limit"]\
    - stats_df["Lower_Limit"])/(6*stats_df["Sigma"]) #プロセス能力指数
    stats_df["Ave+3Sigma"] = stats_df["Average"] + 3*stats_df["Sigma"]#n30平均値+3Sigma
    stats_df["Ave-3Sigma"] = stats_df["Average"] - 3*stats_df["Sigma"]#n30平均値-3Sigma
    stats_df["Difference"] = stats_df["Max_sample"] - stats_df["Min_sample"] #n30最大-最小

    # 統計量(Cpk)行をstat_dfに追加, where構文で条件づけてデータを追加する
    upperCPk = (stats_df["Upper_Limit"] - stats_df["Average"])/(3*stats_df["Sigma"])
    lowerCPk = (stats_df["Average"] - stats_df["Lower_Limit"])/(3*stats_df["Sigma"])
    Cpk = upperCPk.where(upperCPk < lowerCPk, lowerCPk)
    stats_df["Cpk"] = Cpk

    statistics_dfs.append(stats_df)


# In[7]:


#エクセルに転記するdf
pre_export_dfs=[]

for df in statistics_dfs:
    pre_ex_df = pd.DataFrame()
    pre_ex_df = df[["Cpk", "Sigma", "Upper_Limit", "Lower_Limit", "Average",\
"Max_sample", "Min_sample", "Ave+3Sigma", "Ave-3Sigma"]].T

    pre_export_dfs.append(pre_ex_df)


# 各シートの各列ごとにdfを並べ替える.
# ユーザー入力に基づいて作成されるリスト
input_column_name = []

# 各シート名に対してユーザーに入力を促す
for sheet_name in sheet_name_list:
    user_input = input(f"{sheet_name}に対応する文字列を入力してください: ")
    input_column_name.append(user_input)

# ユーザー入力の数をチェック
if len(input_column_name) != len(sheet_name_list):
    print("入力された文字列の数がシートの数と一致しません。プログラムを終了します。")
    sys.exit()

#input_column_name = ["a", "b", "c", "d"] #列名を決める
export_dfs=[]

for column_name in columns_marge:
    ex_df = pd.concat([df[column_name] for df in pre_export_dfs], axis= 1)
    ex_df.columns = [input_column_name[i] for i in range(len(pre_export_dfs))]
    export_dfs.append(ex_df)


#plotするdf
pre_plot_dfs=[]

for df in statistics_dfs:
    pre_plt_df = pd.DataFrame()
    pre_plt_df = df[["Upper_Limit", "Lower_Limit", "Average",\
"Max_sample", "Min_sample", "Ave+3Sigma", "Ave-3Sigma"]].T

    pre_plot_dfs.append(pre_plt_df)

# 各シートの各列ごとにdfを並べ替える.
plot_dfs=[]

for column_name in columns_marge:
    plot_df = pd.concat([df[column_name] for df in pre_plot_dfs], axis= 1)
    plot_df.columns = [input_column_name[i] for i in range(len(pre_plot_dfs))]
    plot_dfs.append(plot_df)


# In[17]:


# オブジェクト指向型でプロット(all dimension plot)
marker_list = ["_", "_", "o", "^", "^", "_", "_"]
color_list =  ["b", "b", "g", "m", "y", "r", '#984ea3']

for df, column_name in zip(plot_dfs, columns_marge):
    fig, ax = plt.subplots(figsize=(3,4))
    for index, markint, colorint in zip(df.index,marker_list, color_list):
        if index == "Upper_Limit" or index == "Lower_Limit":
            ax.plot(df.columns, df.loc[index],color=colorint, marker=markint, label=index)
        else:
            ax.scatter(x=df.columns, y=df.loc[index],color=colorint, marker=markint, label=index)

    ax.set_xlabel('Mold Version')
    ax.set_ylabel('Dimension')
    ax.set_title(column_name)

    ax.set_xticks(range(len(df.columns)))
    ax.set_xticklabels(df.columns)
    ax.set_xlim([-0.5, len(sheet_name_list)-0.5])

    ax.legend()
    ax.grid(True)
    ax.legend(loc="upper left", bbox_to_anchor=(1.02, 1.0,), borderaxespad=0)

    plt.savefig("plt_folder/"+column_name+".png", dpi = 600,bbox_inches='tight')
    plt.close("all")


# In[18]:


# Excelファイルへの書き込み
excel_path = 'output.xlsx'
with pd.ExcelWriter(excel_path) as writer:
    row = 30  # 開始行
    col = 0  # 開始列

    for i, df in enumerate(export_dfs):
        # データフレームをExcelファイルに書き込み
        df.to_excel(writer, startrow=row, startcol=col)

        # 次の列の位置を更新
        col += df.shape[1] + 5  # データフレームの幅 + 5列の間隔

        # 2つごとに行を変更し、列をリセット
        if (i + 1) % 2 == 0:
            row += df.shape[0] + 26  # データフレームの高さ + 10行の間隔
            col = 0

# cmをインチに変換
def cm_to_pixels(cm):
    return cm * 0.393701 * 96

# 画像を挿入する
wb = load_workbook(excel_path)
ws = wb.active

# 画像ファイルのパス
img_paths = ["plt_folder/"+column_name +".png" for column_name in columns_marge ]  # 画像ファイルのパスを指定

# 画像のサイズ（センチメートル）
width_cm = 11.78
height_cm = 9.91

#画像の初期位置
img_row = 8
img_col = 1

for (i, img_path), df in zip(enumerate(img_paths), export_dfs):
    img = Image(img_path)
    # 画像サイズの指定（ピクセルで指定）
    img.width = cm_to_pixels(width_cm)
    img.height = cm_to_pixels(height_cm)

    # 画像を挿入するセルの位置を指定
    cell = ws.cell(row=img_row, column=img_col)  # 例えば、最初のセルに画像を挿入
    ws.add_image(img, cell.coordinate)

    # 次の列の位置を更新
    img_col += df.shape[1] + 5  # データフレームの幅 + 5列の間隔

    # 2つごとに行を変更し、列をリセット
    if (i + 1) % 2 == 0:
        img_row += df.shape[0] + 26  # データフレームの高さ + 10行の間隔
        img_col = 1

wb.save(excel_path)

